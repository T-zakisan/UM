'''
全階層のしおりタイトルとページ取得
分割後のpdfに階層を考慮したしおり情報付与
1-2ページとそれ以降のページを分割
開き方等を設定
セキュリティ付与
'''

from pathlib import Path
import sys
import pikepdf
import fitz
from pikepdf import OutlineItem

OWNER_PASSWORD = "0664721418"
USER_PASSWORD = ""


# バッチ処理用のExcelファイルを読み込み、各行の設定に従ってPDFを分割・出力する
def run_batch_from_excel(excel_path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):  # 2行目から処理
        split_cell, submit_cell, share_cell, filename = row

        if not filename or not Path(filename).is_file():
            print(f" → ファイルが見つかりません: {filename}")
            continue

        input_path = Path(filename)
        name = input_path.stem

        has_split = bool(split_cell and str(split_cell).strip())
        has_submit = bool(submit_cell and str(submit_cell).strip())
        has_share = bool(share_cell and str(share_cell).strip())

        if has_split and has_share and not has_submit:
            print(f"\n■ {filename} → エラー表示（分割＋共有は不可）")
            continue

        print(f"\n■ バッチ処理対象: {filename}")
        result_files = []

        if has_split and not has_submit and not has_share:
            split_pdf(input_path, "しおり", split_mode=True)
            result_files.append(f"{name}_表紙.pdf")
            result_files.append(f"{name}_本体.pdf")

        elif has_submit and not has_share and not has_split:
            split_pdf(input_path, "提出", split_mode=False)
            result_files.append(f"{name}_提出.pdf")

        elif has_share and not has_submit and not has_split:
            split_pdf(input_path, "共有", split_mode=False)
            result_files.append(f"{name}_共有.pdf")

        elif not has_split and not has_submit and not has_share:
            split_pdf(input_path, "しおり", split_mode=False)
            result_files.append(f"{name}_しおり.pdf")

        elif has_submit and has_share and not has_split:
            split_pdf(input_path, "提出", split_mode=False)
            split_pdf(input_path, "共有", split_mode=False)
            result_files.append(f"{name}_提出.pdf")
            result_files.append(f"{name}_共有.pdf")

        else:
            print(" → 処理条件に該当しません。")
            continue

        print("  → 出力ファイル:", "、".join(result_files))



# Excelメニューを作成する関数
def create_excel_menu():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "処理メニュー"
    ws["A1"] = "分割"
    ws["B1"] = "提出"
    ws["C1"] = "共有"
    ws["D1"] = "ファイル名"

    # カレントディレクトリのPDFを列挙
    pdf_files = sorted(Path(".").glob("*.pdf"))
    for idx, pdf_file in enumerate(pdf_files, start=2):  # 2行目から
        ws[f"D{idx}"] = str(pdf_file)

    output_path = Path("処理メニュー.xlsx")
    wb.save(output_path)
    print(f"Excelメニューを作成しました: {output_path.resolve()}")



# 全階層のしおりタイトルとページ取得
def print_bookmarks_with_page(pdf_path: Path) -> list:
    doc = fitz.open(pdf_path)
    toc = doc.get_toc(simple=True)
    print("■ しおり一覧（タイトルとページ番号）:")
    for level, title, page in toc:
        indent = "  " * (level - 1)
        print(f"{indent}- {title}（{page}ページ）")
    doc.close()
    return toc




# 分割後のpdfに階層を考慮したしおり情報付与
def build_outline_hierarchy(toc: list, page_offset: int, target_pdf: pikepdf.Pdf) -> list:
    root = []
    stack = [(0, root)]
    for level, title, page in toc:
        if page <= page_offset:
            continue
        page_in_new_pdf = page - page_offset - 1
        item = OutlineItem(title, page_in_new_pdf)

        while stack and level <= stack[-1][0]:
            stack.pop()
        stack[-1][1].append(item)
        stack.append((level, item.children))
    return root



# 開き方等を設定
def apply_viewer_settings(pdf: pikepdf.Pdf):
    pdf.Root.PageMode = pikepdf.Name("/UseOutlines")
    pdf.Root.PageLayout = pikepdf.Name("/SinglePage")
    first_page = pdf.pages[0]
    pdf.Root.OpenAction = pdf.make_indirect(pikepdf.Dictionary({
        "/Type": pikepdf.Name("/Action"),
        "/S": pikepdf.Name("/GoTo"),
        "/D": [first_page.obj, pikepdf.Name("/Fit")]
    }))



# セキュリティ付与
def encrypt_pdf(input_pdf: Path, output_pdf: Path, mode: str):
        if mode == "提出":
                print_highres = True
                print_lowres = False   # 高解像度を許可
        elif mode == "共有":
                print_highres = False # 低解像度を許可
                print_lowres = True  #
        else:
                raise ValueError("encrypt_pdfは '提出' または '共有' モード専用です")

        with pikepdf.Pdf.open(str(input_pdf)) as pdf:
                pdf.save(
                        str(output_pdf),
                        encryption=pikepdf.Encryption(
                                owner=OWNER_PASSWORD,
                                user=USER_PASSWORD,
                                allow=pikepdf.Permissions(
                                    extract=False,
                                    modify_annotation=False,
                                    modify_assembly=False,
                                    modify_form=False,
                                    modify_other=False,
                                    accessibility=False,
                                    print_highres=print_highres,
                                    print_lowres=print_lowres,
                                )
                        )
                )
        print(f" → セキュリティ付与済（{mode}）: {output_pdf}")



# メタ情報をコピーする関数
def copy_docinfo(src_info: dict, dest_pdf: pikepdf.Pdf):
    for key, value in src_info.items():
        try:
            dest_pdf.docinfo[key] = str(value)
        except Exception as e:
            print(f"  ※ メタ情報 {key} を設定できませんでした: {e}")



# PDFを分割する関数
def split_pdf(input_path: Path, mode: str, split_mode: bool):
    dst_dir = input_path.parent #/ 'dst' / '分割'
    dst_dir.mkdir(parents=True, exist_ok=True)
    name = input_path.stem
    toc = print_bookmarks_with_page(input_path)

    with pikepdf.open(str(input_path)) as pdf:
        total_pages = len(pdf.pages)
        print(f"\n■ ファイル: {input_path.name}（{total_pages}ページ）")
        original_info = dict(pdf.docinfo)
        print("■ メタ情報:")
        for key, value in original_info.items():
            print(f"  {key}: {value}")

        if total_pages < 3 and split_mode:
            print(" → 3ページ未満のためスキップ")
            return

        # === 分割なしモード ===
        if not split_mode:
            # 全体をそのまま本体扱いでコピー
            main_pdf = pikepdf.Pdf.new()
            for i in range(total_pages):
                main_pdf.pages.append(pdf.pages[i])

            with main_pdf.open_outline() as outline:
                nested_items = build_outline_hierarchy(toc, page_offset=0, target_pdf=main_pdf)
                outline.root.extend(nested_items)

            apply_viewer_settings(main_pdf)
            temp_main_path = dst_dir / f"{name}_temp.pdf"
            final_main_path = dst_dir / f"{name}_{mode}.pdf"
            main_pdf.save(str(temp_main_path))
            print(f" → 保存: {temp_main_path}")

        # === 分割ありモード ===
        else:
            # 表紙
            cover_pdf = pikepdf.Pdf.new()
            for i in range(2):
                cover_pdf.pages.append(pdf.pages[i])
            copy_docinfo(original_info, cover_pdf)

            apply_viewer_settings(cover_pdf) # ビューア設定
            with cover_pdf.open_outline() as outline:
                outline.root.append(OutlineItem('表紙', 0)) # しおり追加：表紙
            cover_path = dst_dir / f"{name}_表紙.pdf"
            cover_pdf.save(str(cover_path))
            print(f" → 表紙出力: {cover_path}")


            # 本体
            main_pdf = pikepdf.Pdf.new()
            for i in range(2, total_pages):
                main_pdf.pages.append(pdf.pages[i]) # 2ページ目以降を追加
            copy_docinfo(original_info, main_pdf)   # メタ情報をコピー
            with main_pdf.open_outline() as outline:
                nested_items = build_outline_hierarchy(toc, page_offset=2, target_pdf=main_pdf)
                outline.root.extend(nested_items)

            apply_viewer_settings(main_pdf)
            temp_main_path = dst_dir / f"{name}_temp.pdf"
            final_main_path = dst_dir / f"{name}_本体.pdf"
            main_pdf.save(str(temp_main_path))
            print(f" → 本体保存: {temp_main_path}")

        # セキュリティ適用
        if mode in ("提出", "共有"):
            encrypt_pdf(temp_main_path, final_main_path, mode)
            temp_main_path.unlink()
        else:
            final_main_path.unlink(missing_ok=True)  # 既存ファイルがあれば削除（Python 3.8+）
            temp_main_path.replace(final_main_path)  # 上書き移動

        print(f" → 本体出力（モード={mode}, 分割={split_mode}）: {final_main_path}")




def main():

    if len(sys.argv) == 1:
        print("引数なしで実行されたため、処理メニューExcelを作成します。")
        create_excel_menu()
        return

    if len(sys.argv) == 2 and sys.argv[1].lower().endswith(".xlsx"):
        excel_path = Path(sys.argv[1])
        if not excel_path.exists():
            print(f"{excel_path} が存在しません。")
            return
        print(f"{excel_path.name} に従ってバッチ処理を開始します。")
        run_batch_from_excel(excel_path)
        return


    if "--batch" in sys.argv:
        excel_path = Path("処理メニュー.xlsx")
        if not excel_path.exists():
            print(f"{excel_path} が存在しません。")
            return
        print(f"{excel_path.name} に従ってバッチ処理を開始します。")
        run_batch_from_excel(excel_path)
        return


    if "--mode" not in sys.argv:
        print("使い方: python script.py input.pdf --mode [しおり|提出|共有] [--split yes|no]")
        return

    # 通常の処理
    mode_index = sys.argv.index("--mode")
    input_path = Path(sys.argv[1])
    mode = sys.argv[mode_index + 1]

    if mode not in ("しおり", "提出", "共有"):
        print("モードは 'しおり'、'提出'、'共有' のいずれかを指定してください。")
        return

    if "--split" in sys.argv:
        split_index = sys.argv.index("--split")
        split_value = sys.argv[split_index + 1].lower()
        split_mode = (split_value == "yes")
    else:
        split_mode = True

    split_pdf(input_path, mode, split_mode)



if __name__ == "__main__":
    main()