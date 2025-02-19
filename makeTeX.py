#####################################################################################
# makeTeX.py
#  ExcelファイルからTeX関係の各種ファイルを出力
#   - TeXコマンド(変数として、文章制御等に利用)を***ファイルとして出力
#   - 各種表を*.texファイルとして出力
# [ 2025.02.19 ]
#####################################################################################

import tkinter as tk            # ダイアログ
from tkinter import filedialog  # ダイアログ
import openpyxl                 # Excel操作



#####################################################################################
# create_sty_and_tex_files
#  - Excelのシートによる分岐とシート種ごとの処理関数呼び出し
#  [ 2025.02.19 ]
#####################################################################################
def create_sty_and_tex_files(excel_file_path):

  try:
    # workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook["ReadMe"]

    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
      sheet = workbook[sheet_name]

      # 変数設定ファイル出力
      if "変数" in sheet_name:
        create_variable_definitions(sheet)  # 関数を呼び出す

      # 各表ファイル出力
      elif sheet_index >= 3:  # 4番目以降のシート
      # elif sheet_index == 3:  # 4番目のシート
        create_table_code(sheet)  # 関数を呼び出す

    # workbook.save(excel_file_path)
    workbook.close()
    return 0

  except FileNotFoundError:
    print(f"エラー: ファイル '{excel_file_path}' が見つかりません。")
    return -1
  except Exception as e:
    print(f"エラー: 予期せぬエラーが発生しました: {e}")
    return -1





#####################################################################################
# create_variable_definitions
#  - ExcelのシートからTeXコマンドの定義文をstyファイルで出力
#  [ 2025.02.19 ]
#####################################################################################
def create_variable_definitions( sheet ):
  output_file_path = "myValues.sty"
  with open(output_file_path, 'w', encoding='utf-8') as f:

    # コメント部分
    myStr = f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n" \
            f"% ファイル名：{output_file_path }\n" \
            f"% 内　　　容：変数の設定(TeX制御含む)\n" \
            f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n\n"
    f.write( myStr )

    for idx, row in enumerate( sheet.iter_rows(min_row=2) ):
      col_a_value = row[0].value  # [列A]変数名
      col_b_value = row[1].value  # [列B]設定値
      col_d_value = row[2].value  # [列D]備考
      myVariable = f"\変数：{col_a_value}"

      if col_a_value is not None and col_b_value is not None:
        col_b_value = escape_tex_string(col_b_value) # テキストチェック(エスケープ文字対応)
        f.write(f"\\newcommand{{{myVariable}}}{{{col_b_value}}} % {col_d_value }\n")




#####################################################################################
# escape_tex_string
#  - TeXで安全な文字列に変換する
#  [ 2025.02.19 ]
#####################################################################################
def escape_tex_string(s):
    if isinstance(s, (int, float)):
        s = str(s)
    s = s.replace("\\", "\\\\")
    s = s.replace("{", "\\{")
    s = s.replace("}", "\\}")
    s = s.replace("%", "\\%")
    s = s.replace("$", "\\$")
    s = s.replace("&", "\\&")
    s = s.replace("#", "\\#")
    s = s.replace("~", "\\~")
    s = s.replace("^", "\\^")
    s = s.replace("_", "\\_")
    return s





##############################################################
# create_table_code
#  - Excelの表をTeX(longtable)形式に変換
#  [ 2025.02.18 ]
##############################################################
def create_table_code(sheet):

  tex_file_name = sheet['A1'].value  # A2セルにファイル名(拡張子なし)が記述されている前提
  if not tex_file_name:
    print(f"シート '{sheet['A1'].value}' のセル[A2]にファイル名が記述されていません。")
    exit

  output_file_path = tex_file_name + ".tex" # ファイル名
  with open(output_file_path, 'w', encoding='utf-8') as f:

##########################################################
    # コメント部分出力
    myStr = "%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n" + \
            "% ファイル名：" + output_file_path + "\n" \
            "% 内　　　容：" + sheet['A1'].value + "\n" \
            "%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n\n"
    f.write( myStr ) # テキスト出力

    # 表の設定：インデント含む
    indent = "0mm" # 初期値：インデントなし
    if "あり" in sheet['A2'].value:
      indent = f"\\myLEFTSKIP" # インデントあり時
    myStr = f"\\setlength{{\\tabcolsep}}{{2mm}} % セルの余白\n" \
            f"\\setlength{{\\LTleft}}{{{ indent }}} % 字下げ\n" \
            f"\\setlength{{\\LTright}}{{0pt}} % 右側余白設定\n" \
            f"\\setlength{{\\LTpre}}{{2mm}} % 表前の余白\n" \
            f"\\setlength{{\\LTpost}}{{0mm}} % 表後の余白\n"
    f.write( myStr ) # テキスト出力


    # ヘッダ部
    NumRule = 0 # 罫線(垂線)の本数：表幅補正に使用
    myLength = 0 # 表幅用：インデントの有無による変動算出
    NumColumn = 0 # 最大列数：なにかしらの処理で空白セルが発生しうるため、列数をこれで制限
    myStr = f"\\begin{{longtable}}"
    myStr += "{"

    for row in sheet.iter_rows(min_row=3, max_row=3):
      for idx, cell in enumerate( row[2:] ): #

        # 列数：空白行が入り込みやすいため、ココで制限をかける
        if cell.value is None:
          break
        else:
          NumColumn = idx

        # 背景色
        # tmp = cell.offset(2,0).value # 背景値取得
        # myStr += " >{"
        # if not( tmp is None ): # 背景色（濃度）：入力あり
        #   myStr += f"\\columncolor{{gray!{tmp}}}"


  # >{\columncolor{yellow}\raggedright\arraybackslash}p{40mm}

        # 幅と寄せの設定
        if str(cell.offset(0,0).value).isdigit(): # 幅が数値の場合（*ではない）
          if "S" in cell.offset(1,0).value: # 寄せ：小数点
            tmp = f">{{}}S[table-column-width={cell.offset(0,0).value}mm]" # 寄せ：なし　幅：数値のまま
            myLength += int( cell.offset(0,0).value ) + 4
          else:# 寄せ：S以外（lcr）
            tmp = f"p{{{cell.offset(0,0).value}mm}}" # 幅：数値のまま
            myLength += int( cell.offset(0,0).value ) + 4
        else:
          tmp = f"p{{\\myTableWidth}}" # 修正必要

        if not( "S" in cell.offset(1,0).value ) :
          if "l" in cell.offset(1,0).value: # 寄せ:左
            tmp = f">{{\\raggedright\\arraybackslash}}" + tmp
          elif "c" in cell.offset(1,0).value:# 寄せ:中央
            tmp = f">{{\\centering\\arraybackslash}}" + tmp
          elif "r" in cell.offset(1,0).value:# 寄せ:右
            tmp = f">{{\\raggedleft\\arraybackslash}}" + tmp


        # 罫線
        if cell.offset(1,0).value.replace(" ","")[0:1] == "|": # 左側に罫線
          tmp = f" |{tmp}"
          NumRule += 1
        elif cell.offset(1,0).value.replace(" ","")[-1:] == "|": # 右側に罫線
          tmp = f" {tmp}|"
          NumRule += 1
        myStr = f"{myStr}{tmp} "


        # # 寄せ
        # tmp = cell.offset(1,0).value # 寄せ取得
        # if tmp == "l": # 左寄せ
        #   myStr += f"\\raggedright\\arraybackslash}}" # "}"は"}}""
        # elif tmp == "c" or tmp == "": # 中央寄せ：未記入
        #   myStr += f"\\centering\\arraybackslash}}"
        # elif tmp == "r": # 右寄せ
        #   myStr += f"\\raggedleft\\arraybackslash}}"
        # elif tmp == "S": # 小数点揃え
        #   myStr += f"}}{tmp}"
        # # myStr += "}"

        # # 列幅
        # tmp = f"{cell.offset(0,0).value}" # 列幅取得
        # if tmp.isdigit(): # 数値の場合
        #   if cell.offset(1,0).value == "S":
        #     myStr += f"[table-column-width={tmp}mm]" # 数値をそのまま反映
        #     myLength += int( tmp ) + 4
        #   else:
        #     myStr += f"p{{{tmp}mm}}" # 数値をそのまま反映
        #     myLength += int( tmp ) + 4
        # else:
        #   myStr += f"p{{\\myTableWidth}}" # 修正必要
        #   pass

    myStr += "}\n"

    tmp = f"\\setlength{{\\myTableWidth}}{{\\dimexpr 166mm - {myLength}mm - {indent} - {NumRule}\\arrayrulewidth }}\n"
    f.write( tmp ) # テキス出力
    f.write( myStr ) # テキス出力

    # print( f"{NumColumn} : {sheet['A1'].value}" )
##########################################################
    # リスト＆環境の締め部
    myStr = "" # テキストの一時保管用
    count_row = 0 # 行カウンタ：偶/奇数処理用
    for row in sheet.iter_rows(min_row=6): # 行で繰り返し(7行目以降)
      list = [] # 行項目リスト
      count_row += 1 # 行カウンタの加算
      flag_merge = False # 結合フラグ
      flag_header = False # ヘッダ用フラグ
      flag_hline = "" # 罫線用フラグ（文字列）
      for ii, cell in enumerate( row[0:NumColumn+3] ): # 各行の列で繰り返し(列Aスタート)

        # 罫線チェック
        if ii == 0:
          if cell.value == "改ページ":
            flag_hline = "\\hline \\pagebreak"
          elif cell.value == "太線":
            flag_hline = "\\hline \\hline"
          elif cell.value == "細線":
            flag_hline = "\\hline"
          else:
            flag_hline = ""

        # 行指定の背景色
        if ii == 1: # 列A @ Excel
          if cell.value is None: # 列Bの値が空白の場合
            if ( count_row % 2 ) == 0: # 偶奇数判定
              myStr += f"\\rowcolor{{gray!10}}\t" # 偶数行
            else:
              myStr += f"\\rowcolor{{white}}\t" # 奇数行
          else:
            if str(cell.value).isdigit(): # 数値判定
              myStr += f"\\rowcolor{{gray!{cell.value}}}\t" # 背景色
              if cell.value == 60: # ヘッダフラグ
                if cell.offset(1,0).value != 60: # 次行が60(タイトル：濃背景)の場合
                  flag_header = True #

        # 結合チェック
        if (ii > 1) and (ii<len(row)): # 不要項目除去
          flag_merge = False # 結合フラグ
          for jj in range( len(list) ): # 既存リストで繰り返し
            if cell.value == list[jj][0] and not(list[jj][0] is None): # 結合判定
              flag_merge = True # 結合（隣セルと同値）：あり

          if flag_merge == True: # 結合（隣セルと同値）：あり
            list[jj][1] += 1 # 結合数をカウントアップ
          else: # 結合：なし
            list.append([cell.value,1]) # リストにセル値，数量(1)を追記

      for ii, ll in enumerate( list ): # リスト部の処理
        if ll[1] == 1: # セル結合なし(1=1セル)
          if ll[0] is None: # セル値：なし
            myStr += f" & " # 空白表示
          else: # セル値：あり
            if not( row[1].value is None): # 行背景色：あり
              myStr += f"\\textbf{{{ll[0]}}} & " # 太文字
            else:# 行背景色：なし
              myStr += f"{ll[0]} & " # そのまま表示

        else: # セル結合あり

          if sheet.cell(4,ii+3).value == "S"   : # 寄せがS(小数点)
            align = "c" #
            tmp = f"\\textbf{{{ll[0]}}}"
          elif ii == 0: # 列C
            align = "l" # B列：左寄＆太文字
            tmp = f"\\textbf{{{ll[0]}}}"

          elif not( row[1].value is None) : # 背景色あり(=タイトル)
            align = "l" # 左寄＆太文字
            tmp = f"\\textbf{{{ll[0]}}}"
          else: # 列D以降
            align = "c" # 中央寄せ
            tmp = f"{ll[0]}"
          myStr += f"\\multicolumn{{{ll[1]}}}{{{align}}}{{{tmp}}} & " # セル値

        # 行カウンタの初期か
        if not( row[1].value is None):
          count_row = 0 # 行カウンタの初期化

      myStr = myStr[:-2] + f"\\\\ {flag_hline}\n"
      # if flag_hline in "改ページ":
      #   myStr += f"\\hline\\pagebreak\n" # 罫線＋改ページ
      # else:
      #   myStr +=  # 最後の&除去と改行
      flag_hline = ""

      if flag_header == True:
        flag_header = False
        tmp = myStr
        myStr = f"{tmp} \\endfirsthead\n" \
                f"{tmp} \\endhead\n"

    myStr += f"\n\\end{{longtable}}" # 環境の締め
    f.write( myStr )





# メイン
if __name__ == "__main__":
  root = tk.Tk()
  root.withdraw()
  excel_file_path = filedialog.askopenfilename(title="設定(Excel)ファイルを選択してください", filetypes=[("Excel Files", "*.xlsm")])

  if excel_file_path:
    result = create_sty_and_tex_files(excel_file_path)
    # print( result )
    if result == 0:
      print("処理終了！")
    else:
      print("処理中にエラー発生！")
  else:
    print("ファイルが選択されませんでした。")

