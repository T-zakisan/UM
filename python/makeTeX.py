#####################################################################################
# makeTeX.py
#  2025.03.10
#  - ExcelファイルからTeX関係の各種ファイルを出力
#  - TeXコマンド(変数として、文章制御等に利用)を***.styファイルとして出力
#  - 各種表を*.texファイルとして出力
#####################################################################################

import os
import sys
import glob
from tkinter import messagebox  # ダイアログ
import openpyxl                 # Excel操作
import pathlib                  # パスの処理
import mojimoji                 # 全角半角
import datetime                 # 日付表記の変更


#####################################################################################
# create_sty_and_tex_files
#  - Excelのシートによる分岐とシート種ごとの処理関数呼び出し
# [引数] base_path  Excelファイルの親パス
#####################################################################################
def create_sty_and_tex_files(excel_file_path):

  # 既存の出力ファイル削除
  ff = [ '*.pdf', '*.toc', '*.out', '*.aux', '*.log', 'main.tex' ]
  for pattern in ff:
    for filepath in pathlib.Path('.').glob( pattern ):
      try:
        os.remove(filepath)
      except FileNotFoundError:
        print(f"ファイル {filepath} は存在しません。")
      except Exception as e:
        print(f"ファイル {filepath} の削除中にエラーが発生しました: {e}")

######################
# 1 : × : ReadMe
# 2 : ▲ : 版
# 3 : ▲ : 変数
# 4 : 未 : 安全上の…(b)
# 5 : × : 表リスト
# 6-: ▲ : 各表
######################


  try:
    workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
    # workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook["ReadMe"]

    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
      sheet = workbook[sheet_name]

      # 変数設定ファイル出力
      if "変数" in sheet_name:
        lang = create_variable(sheet, os.path.dirname(excel_file_path))  # 関数を呼び出す

      ## elif "安全上の…b" in sheet_name:
      ##   create_variable_definitions(sheet, os.path.dirname(excel_file_path))  # 関数を呼び出す

      # 各表ファイル出力
      elif sheet_index >= 5:  # 6番目(0スタート)以降のシート
        create_table_code(sheet, pathlib.Path(excel_file_path).parent )  # 関数を呼び出す

    # [版]を後から追記したいので、もうひとサイクル！
    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
      sheet = workbook[sheet_name]

      # 変数設定ファイル出力
      if "版" in sheet_name:
        create_version(sheet, pathlib.Path(excel_file_path).parent, lang )  # 関数を呼び出す

    # workbook.save(excel_file_path)
    workbook.close()



    # makeMain.py　はここにするべき



    return 0




  except FileNotFoundError:
    print(f"エラー: ファイル '{excel_file_path}' が見つかりません。")
    return -1
  except Exception as e:
    print(f"エラー: 予期せぬエラーが発生しました: {e}")
    return -1







#####################################################################################
# create_variable
#  - ExcelのシートからTeXコマンドの定義文をstyファイルで出力
# [引数] sheet      シートオブジェクト
# [引数] base_path  Excelファイルの親パス
#####################################################################################
def create_variable( sheet, base_path ):


  # セル[A3]が空白なら終了：■共通フォルダ対策
  value = sheet.cell(3,1).value
  if value is None or str(value).strip() == "":
    return


  # 変数
  machine_type = ""   # 機種
  machine_model = ""  # 号機
  book_no = ""        # Book No


  # 相対パスのstyファイル(設定パス.sty)出力処理
  base_path = pathlib.Path( base_path )
  output_file_path = base_path.joinpath("設定全体.sty")
  with open(output_file_path, 'w', encoding='utf-8') as ff:

    # ファイルコメント
    myStr = f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n" \
            f"% ファイル名：{output_file_path.name}\n" \
            f"% made by makeTeX.py\n" \
            f"% 内　　　容：日本語名コマンド用コマンド\n" \
            f"%           ：■共通への相対パス（変数・表生成.xlsmを完成のこと！）\n" \
            f"%           ：本文制御用コマンド定義\n" \
            f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n\n"
    ff.write( myStr )

    # 日本語コマンド定義
    myStr = f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n" \
            f"% コマンドの定義チェックからの表示\n" \
            f"%  (これだけ例外的にHANTA.styから独立)\n" \
            f"%  通常の変数（コマンド）と同じように展開され、使える\n" \
            f"%  未定義の変数は未表示\n" \
            f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n" \
            f"\\usepackage{{xparse}} % 限りない引数\n" \
            f"\\usepackage{{expl3}} % スクリプト用\n" \
            f"\\NewExpandableDocumentCommand{{\\変数}}{{m}}{{\\csname 変数:#1\\endcsname}}\n" \
            f"\\NewDocumentCommand{{\\変数設定}}{{mm}}{{\\global\\expandafter\\def\\csname 変数:#1\\endcsname{{#2}}}}\n\n"
    ff.write( myStr )

    myStr = "" # 文字列用変数をリセット
    flag_lang = False # 表示言語用のフラグリセット
    flag_end = False
    for idx, row in enumerate( sheet.iter_rows(min_row=1)):
      col_a_value = row[0].value  # [列A]変数名
      col_b_value = row[1].value  # [列B]設定値
      col_c_value = row[2].value  # [列C]備考

      # １行目：パス
      if idx == 0 :
        col_b_value = col_b_value.replace( "/", "\\" )
        each_dir = str(base_path).split("\\")  # /区切り
        for ii, dir_name in enumerate( each_dir ):
          if dir_name == "●機種":
            dir_depth = len(each_dir)-ii
            break
        relative_path = "../" * dir_depth# 相対パス生成
        myStr += f"\\変数設定{{■共通パス}}{{{relative_path}■共通}} % 共通パーツにアクセスするためのパス\n\n"
        myStr += f"% 主に，表紙 ＆ PDFの文書プロパティ\n"

      # ３行目以降 & 文字色が青
      if idx >= 2 and row[0].font.color.rgb == "FF0070C0" : # 色縛り：青
        # 表示言語による機種名の全角/半角切り替え
        ## 表示言語の行でフラグセット
        if (col_a_value == "表示言語") and (col_b_value == "Jpn") :
          flag_lang  = True

        ##################################################
        ## 表示言語による機種名の全/半角切り替え
        if flag_lang  == True:
          if  col_a_value == "機種名" or \
              col_a_value == "機種" :
            col_b_value = mojimoji.han_to_zen( col_b_value ) # 日本語表記[Jpn]：全角化
        else:
          pass
          # col_b_value = mojimoji.han_to_zen( col_b_value ) # 基本的には半角表記

        ##################################################
        if col_a_value != "号機":
          # 号機以外の青文字
          myStr += f"\\変数設定{{{col_a_value}}}{{{col_b_value}}} % {col_c_value }\n"
        else:
          # 号機
          if row[0].offset(1,0).value == "号機": # 次行の列Aの値が号機か確認
            # 次行も号機：表形式で表示させるためのデータ生成
            ## 号機の範囲取得
            result = {} # 表用データ格納用
            types = [] # 号機一時格納配列
            num = 0 # "号機"の行数カウント
            while 1 : # 無限ループ
              if row[0].offset(num,0).value == "号機": # 号機縛り
                types.append( row[1].offset(num,0).value ) # 型式号機
                num += 1
              else:
                break # "号機"でない行となったら終了
            for type in types: # 取得した号機sで繰り返し
              if type.find("；") == -1:
                result.setdefault("辞書",[]).append(type) # ；区切以降を辞書登録
              else:
                result.setdefault(type[:type.find("；")],[]).append(type[type.find("；")+1:]) # ；区切以降を辞書登録

            flag_end = True # 終了フラグ
          else:
            # 次行号機ではない
            myStr += f"\\変数設定{{{col_a_value}}}{{{col_b_value}}} % {col_c_value }\n"


        # ここから"前後左右の方向"用の値取得
        if col_a_value == "機種名":
          machine_type = col_b_value
        if col_a_value == "機種" :
          machine_model = col_b_value
        if col_a_value == "BookNo":
          book_no = col_b_value

      if flag_end == True:
        break

    if flag_end == True:
      # 号機を表形式表示するためのデータ整形
      # print( f"result : {result}" )
      # input()
      keys_str = ' & '.join(result.keys()) # キーを結合
      values_str = ''
      for ii in range( max(map(len, result.values())) ):
        temp_values = []
        for key in result:
          if ii < len(result[key]):
            temp_values.append(result[key][ii])
          else:
            temp_values.append('')
        values_str += ' & '.join(temp_values) + ' \\\\'
      result_str = keys_str + ' \\\\' + values_str

      if result_str.find("辞書") != -1: # 複数行、小型用ではない場合
        result_str = f"{result_str[result_str.find(";")+1:-1].replace("\\\\","・")}以降"
        myStr += f"\\変数設定{{号機}}{{{result_str}}} % 複数号機\n"
      elif  result_str.find("辞書") == -1: # 複数行、小型(;区切り)の場合
        myStr += f"\\変数設定{{号機s}}{{{result_str}}} % 複数号機\n"
        myStr += f"\\変数設定{{号機数}}{{{len(result)}}} % 複数号機\n"

    myStr += f"\n\n% 版/発行日\n"
    ff.write( myStr ) # テキスト書き込み





  for idx, row in enumerate( sheet.iter_rows(min_row=1)):

    if row[0].value == "前後左右方向の表現":
      filename = base_path.joinpath( "011_前後左右方向の表現.tex" )
      if row[1].value is None:
        filename.unlink() # 削除

      elif row[1].value == "あり":
        # 前後左右方向の表現.tex 生成
        if filename.exists :
          with open(filename, 'w', encoding='utf-8') as ff:
            myStr = f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n" \
                    f"% 011_前後左右方向の表現.tex\n" \
                    f"% made by makeTeX.py\n" \
                    f"% 機　種：{machine_type}\n" \
                    f"% 型　式：{machine_model}\n" \
                    f"% BookNo：{book_no}\n" \
                    f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n\n\n" \
                    f"% % \\clearpage % 改ページ：図、表など出力してから改ページ\n" \
                    f"% % \\vspace{{11pt}}\n" \
                    f"% % \\vspace{{5.0mm}}\n" \
                    f"\\IfEq{{\\変数{{前後左右方向の表現}}}}{{あり}}{{%\n" \
                    f"  \\IfEq{{\\変数{{表示言語}}}}{{Jpn}}%\n" \
                    f"    {{\\subsectionSP{{前後左右方向の表現}}}}%\n" \
                    f"    {{\\subsectionSP{{Representation of front/back/left/right}}}}%\n" \
                    f"}}{{}}\n\n\n" \
                    f"% ※'ファイル名'が'方向.pdf'ではない場合は，状況に応じて変更すること！\n" \
                    f"% ■前後左右方向の画像作成ルール■■■■■■■■■■■■■\n" \
                    f"% 1.画像ファイル名\n" \
                    f"%     各機種フォルダの'図/方向.pdf'とすること！\n" \
                    f"%       →以下コード変更の工数削減\n" \
                    f"% 2.画像サイズ\n" \
                    f"%     幅×高＝mm×mmの寸法に収まる画像を作成すること！\n" \
                    f"%       →縮尺変更時に文字が拡大/縮小防止のため\n" \
                    f"% 3.画像の向き\n" \
                    f"%     掲載時の向きで作成すること！\n" \
                    f"%       →以下コード変更の工数削減\n" \
                    f"%       ※Inkscape : 回転が必要な場合は，右下のR値を90/-90とし画像を事前に逆回転させておくこと！\n" \
                    F"% ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■\n" \
                    f"\\begin{{figure}}[h]\n" \
                    f"\\centering\n" \
                    f"\\includepraphic[scale=1.0, angle=0, keepaspectratio]{{./図/方向.pdf}} % 画像サイズのまま，回転なし　※原則コレを使用！\n" \
                    f"%\\includepraphic[scale=1.0, angle=90, keepaspectratio]{{./図/方向.pdf}} % 画像サイズのまま，90度回転(CCW)\n" \
                    f"%\\includepraphic[width=\\textwidth, angle=0, keepaspectratio]{{./図/方向.pdf}} % 用紙幅にフィット，回転なし\n" \
                    f"%\\includepraphic[height=\\textwidth, angle=90, keepaspectratio]{{./図/方向.pdf}}% 用紙幅にフィット後に90度回転(CCW)\n" \
                    f"\\end{{figure}}\n"
            ff.write( myStr ) # テキスト書き込み
        break





  # 自作変数のstyファイル(設定変数.sty)出力処理
  output_file_path = base_path.joinpath( "設定変数.sty")
  with open(output_file_path, 'w', encoding='utf-8') as ff:

    # コメント部分
    myStr = f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n" \
            f"% ファイル名：{output_file_path.name}\n" \
            f"% made by makeTeX.py\n" \
            f"% 内　　　容：自作変数の設定(TeX制御含む、変数・表生成.xlsmを完成のこと！）\n" \
            f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n\n"
    ff.write( myStr )

    for idx, row in enumerate( sheet.iter_rows(min_row=3) ):
      col_a_value = row[0].value  # [列A]変数名
      col_b_value = row[1].value  # [列B]設定値
      col_c_value = row[2].value  # [列C]備考
      if (col_a_value is not None) and (col_b_value is not None) and row[0].font.color.rgb != "FF0070C0": # 色縛：青以外
        ff.write(f"\\変数設定{{{col_a_value}}}{{{col_b_value}}} % {col_c_value }\n")

  return flag_lang  # 表示言語フラグ




#####################################################################################
# escape_tex_string
#  - TeXで安全な文字列に変換する
# [引数] s 文字列
#####################################################################################
def escape_tex_string(s):
    if isinstance(s, (int, float)):
        s = str(s)
    s = s.replace("\\", "\\\\")
    s = s.replace("{", "\\{")
    s = s.replace("}", "\\}")
    s = s.replace("%", "\\%")
    s = s.replace("$", "\\$")
    s = s.replace("&", "\\&")
    s = s.replace("#", "\\#")
    s = s.replace("~", "\\~")
    s = s.replace("^", "\\^")
    s = s.replace("_", "\\_")
    return s



# メイン
if __name__ == "__main__":

  ret = messagebox.askyesno('確認', '［は　い］：誤動作⇒終了！ \n\n［いいえ］：■共通の表作成')   #「はい」、「いいえ」を選択
  if not ret:
    script_dir = pathlib.Path( sys.argv[0] ).parent.parent
    excel_file_path = script_dir.joinpath("変数・表生成.xlsm")

