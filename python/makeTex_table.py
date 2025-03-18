##############################################################
# createTable.py
#  - Excelの表をTeX(longtable)形式に変換
# [引数] sheet      シートオブジェクト
# [引数] base_path  Excelファイルの親パス
##############################################################

import os
import sys
import glob
from tkinter import messagebox  # ダイアログ
import openpyxl                 # Excel操作
import pathlib                  # パスの処理
import mojimoji                 # 全角半角
import datetime                 # 日付表記の変更


def create_table_code(sheet, base_path):

  tex_file_name = sheet['A1'].value  # A2セルにファイル名(拡張子なし)が記述されている前提
  if not tex_file_name:
    print(f"シート '{sheet['A1'].value}' のセル[A2]にファイル名が記述されていません。")
    exit

  base_path = pathlib.Path( base_path )
  output_dir = base_path.joinpath("表")
  if not os.path.exists( output_dir ):
    os.makedirs( output_dir ) # なければディレクトリ生成

  output_file_path = output_dir.joinpath("{tex_file_name}.tex") # ファイル名
  with open(output_file_path, 'w', encoding='utf-8') as f:

##########################################################
    # コメント部分出力
    myStr = f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n" \
            f"% ファイル名：{output_file_path.stem}\n" \
            f"% made by makeTeX.py\n" \
            f"% 内　　　容：{sheet['A1'].value}\n" \
            f"%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%\n\n"
    f.write( myStr ) # テキスト出力

    # 表の設定：インデント含む
    indent = "0mm" # 初期値：インデントなし
    if "あり" in sheet['A2'].value:
      indent = f"\\myLEFTSKIP" # インデントあり時
    myStr = f"\\setlength{{\\tabcolsep}}{{2mm}} % セルの余白\n" \
            f"\\setlength{{\\LTleft}}{{{ indent }}} % 字下げ\n" \
            f"\\setlength{{\\LTright}}{{0pt}} % 右側余白設定\n" \
            f"\\setlength{{\\LTpre}}{{2mm}} % 表前の余白\n" \
            f"\\setlength{{\\LTpost}}{{0mm}} % 表後の余白\n\n\n"
    # f.write( myStr ) # テキスト出力


    # ヘッダ部
    NumRule = 0 # 罫線(垂線)の本数：表幅補正に使用
    myLength = 0 # 表幅用：インデントの有無による変動算出
    NumColumn = 0 # 最大列数：なにかしらの処理で空白セルが発生しうるため、列数をこれで制限


    if "ボルト締付" in sheet['A1'].value:
      myStr = f"{{\\ttfamily % [ボルト締付トルク] など等幅フォントを使用する場合は、先頭の[%]を省く\n\n"

    myStr += f"\\begin{{longtable}}"
    myStr += "{"

    for row in sheet.iter_rows(min_row=3, max_row=3):
      for idx, cell in enumerate( row[2:] ): #

        # 列数：空白行が入り込みやすいため、ココで制限をかける
        if cell.value is None:
          break
        else:
          NumColumn = idx


        # 幅と寄せの設定
        if str(cell.offset(0,0).value).isdigit(): # 幅が数値の場合（*ではない）
          if "S" in cell.offset(1,0).value: # 寄せ：小数点
            tmp = f">{{}}S[table-column-width={cell.offset(0,0).value}mm]" # 寄せ：なし　幅：数値のまま
            myLength += int( cell.offset(0,0).value ) + 4
          else:# 寄せ：S以外（lcr）
            tmp = f"p{{{cell.offset(0,0).value}mm}}" # 幅：数値のまま
            myLength += int( cell.offset(0,0).value ) + 4
        else:
          tmp = f"p{{\\myTableWidth}}" # 修正必要

        if not( "S" in cell.offset(1,0).value ) :
          if "l" in cell.offset(1,0).value: # 寄せ:左
            tmp = f">{{\\raggedright\\arraybackslash}}" + tmp
          elif "c" in cell.offset(1,0).value:# 寄せ:中央
            tmp = f">{{\\centering\\arraybackslash}}" + tmp
          elif "r" in cell.offset(1,0).value:# 寄せ:右
            tmp = f">{{\\raggedleft\\arraybackslash}}" + tmp


        # 罫線
        if cell.offset(1,0).value.replace(" ","")[0:1] == "|": # 左側に罫線
          tmp = f" |{tmp}"
          NumRule += 1 # 罫線数を加算
        if cell.offset(1,0).value.replace(" ","")[-1:] == "|": # 右側に罫線
          tmp = f" {tmp}|"
          NumRule += 1 # 罫線数を加算
        myStr = f"{myStr}{tmp} "

    myStr += "}\n"

    tmp = f"\\setlength{{\\myTableWidth}}{{\\dimexpr 166mm - {myLength}mm - {indent} - {NumRule}\\arrayrulewidth }}\n"
    f.write( tmp ) # テキス出力
    f.write( myStr )

    # print( f"{NumColumn} : {sheet['A1'].value}" )
##########################################################
    # リスト＆環境の締め部
    myStr = "" # テキストの一時保管用
    count_row = 0 # 行カウンタ：偶/奇数処理用
    for row in sheet.iter_rows(min_row=6): # 行で繰り返し(7行目以降)
      list = [] # 行項目リスト
      count_row += 1 # 行カウンタの加算
      flag_merge = False # 結合フラグ
      flag_header = False # ヘッダ用フラグ
      flag_hline = "" # 罫線用フラグ（文字列）
      for ii, cell in enumerate( row[0:NumColumn+3] ): # 各行の列で繰り返し(列Aスタート)

        # 罫線チェック
        if ii == 0:
          if cell.value == "改ページ":
            flag_hline = "\\hline \\pagebreak"
          elif cell.value == "太線":
            flag_hline = "\\hline \\hline"
          elif cell.value == "細線":
            flag_hline = "\\hline"
          else:
            flag_hline = ""

        # 行指定の背景色
        if ii == 1: # 列A @ Excel
          if cell.value is None: # 列Bの値が空白の場合
            if ( count_row % 2 ) == 0: # 偶奇数判定
              myStr += f"\\rowcolor{{gray!10}}\t" # 偶数行
            else:
              myStr += f"\\rowcolor{{white}}\t" # 奇数行
          else:
            if str(cell.value).isdigit(): # 数値判定
              myStr += f"\\rowcolor{{gray!{cell.value}}}\t" # 背景色
              if cell.value == 60: # ヘッダフラグ
                if cell.offset(1,0).value != 60: # 次行が60(タイトル：濃背景)の場合
                  flag_header = True #

        # 結合チェック
        if (ii > 1) and (ii<len(row)): # 不要項目除去
          flag_merge = False # 結合フラグ
          for jj in range( len(list) ): # 既存リストで繰り返し
            if cell.value == list[jj][0] and not(list[jj][0] is None): # 結合判定
              flag_merge = True # 結合（隣セルと同値）：あり

          if flag_merge == True: # 結合（隣セルと同値）：あり
            list[jj][1] += 1 # 結合数をカウントアップ
          else: # 結合：なし
            list.append([cell.value,1]) # リストにセル値，数量(1)を追記

      for ii, ll in enumerate( list ): # リスト部の処理
        if ll[1] == 1: # セル結合なし(1=1セル)
          if ll[0] is None: # セル値：なし
            myStr += f" & " # 空白表示
          else: # セル値：あり
            if not( row[1].value is None): # 行背景色：あり
              myStr += f"\\textbf{{{ll[0]}}} & " # 太文字
            else:# 行背景色：なし
              myStr += f"{ll[0]} & " # そのまま表示

        else: # セル結合あり

          if sheet.cell(4,ii+3).value == "S"   : # 寄せがS(小数点)
            align = "c" #
            tmp = f"\\textbf{{{ll[0]}}}"
          elif ii == 0: # 列C
            align = "l" # B列：左寄＆太文字
            tmp = f"\\textbf{{{ll[0]}}}"

          elif not( row[1].value is None) : # 背景色あり(=タイトル)
            align = "l" # 左寄＆太文字
            tmp = f"\\textbf{{{ll[0]}}}"
          else: # 列D以降
            align = "c" # 中央寄せ
            tmp = f"{ll[0]}"
          myStr += f"\\multicolumn{{{ll[1]}}}{{{align}}}{{{tmp}}} & " # セル値

        # 行カウンタの初期か
        if not( row[1].value is None):
          count_row = 0 # 行カウンタの初期化

      myStr = myStr[:-2] + f"\\\\ {flag_hline}\n"
      # if flag_hline in "改ページ":
      #   myStr += f"\\hline\\pagebreak\n" # 罫線＋改ページ
      # else:
      #   myStr +=  # 最後の&除去と改行
      flag_hline = ""

      if flag_header == True:
        flag_header = False
        tmp = myStr
        myStr = f"{tmp} \\endfirsthead\n" \
                f"{tmp} \\endhead\n"

    myStr += f"\n\\end{{longtable}}\n\n" # 環境の締め

    if "ボルト締付" in sheet['A1'].value:
      myStr += f"}} % [ボルト締付トルク] など等幅フォントを使用する場合は、先頭の[%]を省く\n\n"

    f.write( myStr )

