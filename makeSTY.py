import tkinter as tk
from tkinter import filedialog
import openpyxl

# 
import makeValuable
import makeTable
import utils


def create_sty_and_tex_files(excel_file_path):

  try:
    workbook = openpyxl.load_workbook(excel_file_path, keep_vba=True)
    sheet = workbook["ReadMe"]

    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
      sheet = workbook[sheet_name]

      # 変数設定ファイル出力
      if "変数" in sheet_name:
        makeValuable.create_variable_definitions(sheet)  # 関数を呼び出す

      # 各表ファイル出力
      # elif sheet_index >= 3:  # 4番目以降のシート
      elif sheet_index == 3:  # 4番目のシート
        makeTable.create_table_code(sheet)  # 関数を呼び出す

    workbook.save(excel_file_path)

  except FileNotFoundError:
    print(f"エラー: ファイル '{excel_file_path}' が見つかりません。")
    exit
  except Exception as e:
    print(f"エラー: 予期せぬエラーが発生しました: {e}")
    exit

from openpyxl import load_workbook
# メイン
if __name__ == "__main__":
  root = tk.Tk()
  root.withdraw()
  excel_file_path = filedialog.askopenfilename(title="設定(Excel)ファイルを選択してください", filetypes=[("Excel Files", "*.xlsm")])

  if excel_file_path:
    create_sty_and_tex_files(excel_file_path)
    print("styファイルとtexファイルが作成されました。")
  else:
    print("ファイルが選択されませんでした。")
